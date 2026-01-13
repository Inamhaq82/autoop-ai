from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook, load_workbook

from autoops.products.lead_followup_v1.contracts import Lead, UrgencyLabel

SHEET = "Leads"
HEADERS = ["received_at", "lead_id", "from_address", "subject", "urgency", "status", "raw_ref"]

def ensure_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(HEADERS)
    wb.save(path)
def get_existing_lead_ids(path: Path) -> set[str]:
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET]

    # Find the "lead_id" column index
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        lead_id_col = header_row.index("lead_id") + 1  # 1-based
    except ValueError:
        # If headers are missing/messed up, treat as empty
        return set()

    ids: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        # row tuple is 0-based; convert to col index
        lead_id = row[lead_id_col - 1]
        if isinstance(lead_id, str) and lead_id.strip():
            ids.add(lead_id.strip())
    return ids

def append_lead(path: Path, lead: Lead, urgency: UrgencyLabel, status: str = "RECEIVED") -> bool:
    ensure_workbook(path)
    existing = get_existing_lead_ids(path)
    if lead.lead_id in existing:
        return False

    wb = load_workbook(path)
    ws = wb[SHEET]
    ws.append([
        lead.received_at,
        lead.lead_id,
        lead.from_address,
        lead.subject,
        urgency.value,
        status,
        lead.raw_ref,
    ])
    wb.save(path)
    return True
